"""
LLM Coding Script — Gates Foundation / Norman Lear Center Manfluencer Study
Uses the authenticated Claude Code CLI as the LLM backend (no separate API key needed).
Codes 200 Nigeria + 200 Kenya rows for BOTH content analysis and audience analysis.

Run:
    python scripts/llm_code_all.py --mode content
    python scripts/llm_code_all.py --mode audience
    python scripts/llm_code_all.py --mode all   (default)
"""

import argparse
import asyncio
import json
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT           = Path("/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project")
CONTENT_MASTER = ROOT / "Codebooks/Human Codebooks/Master Human Content Codebook.xlsx"
AUDIENCE_MASTER= ROOT / "Codebooks/Human Codebooks/Master Human Audience Codebook.xlsx"
LLM_DIR        = ROOT / "Codebooks/LLM Codebook"
CACHE_DIR      = ROOT / "temp"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
LLM_DIR.mkdir(parents=True, exist_ok=True)

CLAUDE_BIN = ("/Users/sushildalavi/Library/Application Support/Claude/"
              "claude-code/2.1.138/claude.app/Contents/MacOS/claude")

# ── Config ─────────────────────────────────────────────────────────────────────
MODEL_TAG   = "claude-code-v1"   # cache namespace — change to bust cache
CONCURRENCY = 4
MAX_RETRIES = 3

# ── Excel styling ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL     = PatternFill("solid", fgColor="EBF0FA")

# ══════════════════════════════════════════════════════════════════════════════
#  SHARED UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def load_cache(path: Path) -> pd.DataFrame:
    if path.exists():
        df = pd.read_parquet(path)
        print(f"  Cache loaded: {len(df)} rows ({path.name})")
        return df
    print(f"  No cache found — starting fresh ({path.name})")
    return pd.DataFrame()


def save_cache(df: pd.DataFrame, path: Path):
    df.to_parquet(path, index=False)


def cache_key(item_id: str, country: str, analysis: str) -> str:
    return f"{analysis}::{country}::{item_id}::{MODEL_TAG}"


async def call_claude(prompt: str, item_id: str, attempt_no: int = 0) -> dict:
    """Call the Claude Code CLI via subprocess and return parsed JSON."""
    proc = await asyncio.create_subprocess_exec(
        CLAUDE_BIN, "--print",
        stdin=asyncio.subprocess.PIPE,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await proc.communicate(input=prompt.encode("utf-8"))
    raw = stdout.decode("utf-8", errors="replace").strip()

    # Extract JSON from response (model sometimes wraps in markdown)
    json_match = re.search(r'\{[\s\S]*\}', raw)
    if not json_match:
        raise ValueError(f"No JSON found in response for {item_id}. Raw: {raw[:200]}")
    return json.loads(json_match.group(0))


async def call_with_retry(prompt: str, item_id: str) -> dict:
    for attempt in range(MAX_RETRIES):
        try:
            result = await call_claude(prompt, item_id, attempt)
            return result
        except (json.JSONDecodeError, ValueError) as e:
            wait = 2 ** attempt
            print(f"    JSON error for {item_id} (attempt {attempt+1}): {e} — retrying in {wait}s")
            await asyncio.sleep(wait)
        except Exception as e:
            wait = 2 ** attempt
            print(f"    Error for {item_id} (attempt {attempt+1}): {e} — retrying in {wait}s")
            await asyncio.sleep(wait)
    raise RuntimeError(f"Failed to code {item_id} after {MAX_RETRIES} attempts")


def write_excel_file(path: Path, sheets: dict, run_date: str, methodology_text: list):
    """
    sheets: dict of {sheet_name: pd.DataFrame}
    methodology_text: list of [row, ...] for methodology sheet
    """
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, df in sheets.items():
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = sheet_name

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 40
        for row_idx in range(2, len(df) + 2):
            ws.row_dimensions[row_idx].height = 60

        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for _, row in df.iterrows():
                val = str(row.iloc[col_idx - 1])
                max_len = max(max_len, min(len(val.split("\n")[0]), 80))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    # Methodology sheet
    ws_meth = wb.create_sheet("Methodology")
    for r_idx, row_data in enumerate(methodology_text, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_meth.cell(row=r_idx, column=c_idx, value=str(val))
            if r_idx == 1:
                cell.font = Font(bold=True, size=13)
            cell.alignment = Alignment(wrap_text=True)
    ws_meth.column_dimensions["A"].width = 30
    ws_meth.column_dimensions["B"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  Saved: {path}")


def clean_list_field(val, valid_set):
    if not isinstance(val, list):
        val = []
    return [v for v in val if v in valid_set]


def clean_single_field(val, valid_set, default=""):
    if val in valid_set:
        return val
    return default


def clean_open_text(val) -> str:
    return str(val or "").strip()


def ensure_substantive(text: str, fallback: str) -> str:
    t = clean_open_text(text)
    if len(t.split()) >= 5:
        return t
    return fallback


# ══════════════════════════════════════════════════════════════════════════════
#  CONTENT ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

CONTENT_OUTPUT_COLUMNS = [
    "Content ID", "Context", "Content Text / Description",
    "Q1. Attention-getter (Yes/No)", "Q1a. Attention-getting strategies", "Q1b. Other strategy",
    "Q2. Primary topic(s)", "Q2a. Other topic",
    "Q3. Type of content", "Q3a. Other content type",
    "Q4. Addresses masculinity / gender norms",
    "Q5. Type of masculinity / gender norms",
    "Q6. Addresses what men should do",
    "Q7. What men do or need to do", "Q7a. Other directive",
    "Q8. Problem identified", "Q8a. Other problem",
    "Q9. Solution proposed", "Q9a. Other solution",
    "Q10. Communication mode", "Q10a. Other communication mode",
    "Q11. Audience needs",
    "Q12. How claims are supported", "Q12a. Other claim support",
    "Q13. How claims are justified", "Q13a. Other justification",
    "Q14. Sentiment toward men",
    "Q15. Sentiment toward women",
    "Q16. Sentiment toward traditional gender norms",
    "Q17. Fear or threat used",
    "Q18. Calls to action present",
    "Q18a. Types of calls to action", "Q18b. Other call to action",
]

CONTENT_VALID = {
    "q1":   {"Yes", "No"},
    "q1a":  {"Compelling question", "Use of all CAPS", "Humor or sarcasm",
             "Shares something violent or gross", "Shares something sexual",
             "Shares something surprising", "Uses a news headline or social media trend as opener",
             "Interesting visual or meme", "Other"},
    "q2":   {"Dating/marriage", "Friends/socializing", "Family/children", "Money/status",
             "Fitness/self-improvement", "Mental health", "Gender issues, e.g. equality",
             "Social issues, e.g. corruption", "Religion/morality", "Gaming/technology", "Other"},
    "q3":   {"Interview/conversational content", "Motivational/self-help content",
             "Commentary/reaction content", "Other"},
    "q4":   {"Yes, explicitly", "Yes, implicitly", "No"},
    "q5":   {"More regressive/traditional/restrictive", "More progressive/equitable/expansive",
             "Mixed/unclear", "Does not address masculinity or gender norms"},
    "q6":   {"Yes", "No", "Unclear", "Not applicable"},
    "q7":   {"Men need to dominate/lead", "Men need to provide/succeed",
             "Men are disadvantaged/victims", "Men need to improve themselves",
             "Men need to be fully self-reliant", "Men need to be emotionally open",
             "Men need to not show emotions", "Men need to be equal partners",
             "Mixed/unclear", "Other", "Not applicable"},
    "q8":   {"Kenyan or Nigerian political/social problems", "Global political/social/cultural problems",
             "Western political/social influence", "Women/feminism", "Men's behavior",
             "Economic/status pressure", "Mental health/emotional struggle",
             "No clear problem is identified", "Other"},
    "q9":   {"Social or political change", "Assert dominance/control", "More wealth/status",
             "More self-discipline/fitness", "More emotional growth/healing",
             "More equality/respect for men", "More equality/respect for women",
             "Building community", "No clear solution", "Other"},
    "q10":  {"Advice/instruction", "Personal story", "Commentary/opinion", "Debate/argument",
             "Humor/satire", "Motivational speech", "News/telling facts", "Other"},
    "q11":  {"Entertainment/escapism", "Information seeking", "Connection/social interaction",
             "Self expression/identity construction", "Status seeking",
             "Documentation of events", "None of these apply"},
    "q12":  {"Generalizations about men/women", "Personal experience", "Stories about men/women",
             "Cultural/social observations", "Facts/statistics", "Moral/religious claims",
             "Mixed", "No support", "Other"},
    "q13":  {"No justification", "Anecdotal examples", "Presented as common sense",
             "References data", "References religion/tradition",
             "References external sources, such as other influencers", "Other"},
    "q14":  {"Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"},
    "q15":  {"Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"},
    "q16":  {"Negative", "Positive", "Mixed", "Neutral", "Unclear", "Not mentioned"},
    "q17":  {"Yes", "Somewhat", "No"},
    "q18":  {"Yes", "No"},
    "q18a": {"Calls for audience to like the content", "Calls for audience to share the content",
             "Calls for audience to follow the speaker on social media",
             "Calls for men to follow more traditional gender norms",
             "Calls for men to follow more equitable gender norms",
             "Calls for women to follow more traditional gender norms",
             "Calls for women to follow more equitable gender norms",
             "Calls for politicians or social figures to do something",
             "Calls for audience to vote in a different way", "Other"},
}

THEME_TAXONOMY = """=== RESEARCH THEME TAXONOMY (background context — not a question to answer) ===
Use this taxonomy for domain context while answering Q1–Q18b. Do NOT produce theme codes as output.

1. Authority and Submission — explicit hierarchy, headship, obedience, control, male leadership over women/family, surrendering independence
2. Male Victimhood — men/boys framed as harmed, exploited, ignored, falsely accused, abused, or disadvantaged. Central claim: men are harmed.
3. Gender Grievance — generalized distrust, resentment, or hostility toward women, feminists, or gender equality. Central claim: women/feminism are corrupt or dangerous. (Separate from Male Victimhood: here the THREAT is women/feminism, not harm to men)
4. Sexual Morality — cheating, body count, abortion, pornography, purity, double standards, sexual respectability
5. Relationship Tactics — tactical dating/relationship advice: scarcity, pursuit, withdrawal, dominance, masculine frame
6. Provider and Status — money, career, income, status, economic value as proof of manhood or relationship worth
7. Male Accountability — men must change, hold other men accountable, challenge harmful norms
8. Egalitarian Partnership — mutual respect, shared parenting, allyship, reciprocity, non-hierarchical partnership
9. Gender-Based Violence and Consent — rape, consent, sexual assault, abuse, victim-blaming, femicide, FGM, prevention
10. Trauma and Mental Health — trauma, depression, grief, shame, emotional suppression, healing, therapy, vulnerability
11. Self-Discipline and Self-Mastery — discipline, restraint, fitness, hustle, personal growth, self-improvement (code regardless of whether the framing is regressive or progressive)
12. Marriage and Family — marriage, divorce, fatherhood, parenting, family stability (if submission/headship dominates → use Authority and Submission instead)
13. Faith and Moral Repair — God, prayer, scripture, spiritual healing tied to masculinity/gender/family/relationships (only when gender/masculinity is a primary frame)
Unclear — uncodable, off-topic, low-signal, no clear gender/masculinity theme

Tie-breaker rules:
- "Submit/obey/surrender/head of household" → Authority and Submission (not Marriage and Family)
- Dating tactics/scarcity/frame → Relationship Tactics (not Provider and Status)
- Men harmed/ignored → Male Victimhood (not Gender Grievance)
- Women/feminism framed as threat/corrupt → Gender Grievance (not Male Victimhood)
- False accusations in GBV context → GBV and Consent; as general grievance → Male Victimhood or Gender Grievance
- Self-discipline even if used regressively → Self-Discipline and Self-Mastery
- Faith as primary moral frame for gender/healing → Faith and Moral Repair
"""

CONTENT_SYSTEM = """You are a precise qualitative research coder for a study on masculinity-focused social media influencers in Nigeria and Kenya. Your task is to code content items using a structured codebook.

Coding rules:
- Answer ONLY using the exact options listed for each question.
- For multi-select questions, return a JSON array of strings.
- For single-select questions, return a single string.
- For open-text questions, return a string (empty string if not applicable).
- Return ONLY valid JSON — no markdown fences, no commentary.
- If you select "Other", always provide substantive detail in the matching open-text field (minimum 8 words). Never leave an Other open-text field blank.
- Be analytically thorough — do not default to "No" without real consideration.
- Read the full text carefully before coding."""


def build_content_prompt(content_id: str, country: str, context: str, text: str) -> str:
    country_note = (
        "Nigeria context: Pidgin/Yoruba/Igbo/Hausa terms may appear. "
        "Faith framing is common (Christian + Islamic). Creators: Agba John Doe, Banky Wellington, "
        "Deyemi Okanlawon, Ebuka Obi-Uchendu, Shola, Wizarab."
        if country == "Nigeria" else
        "Kenya context: Sheng/Swahili terms may appear. Andrew Kibe/Jagero is a regressive manosphere "
        "influencer. Rixpoet/Onyango is a progressive mental health advocate. Philip Karanja focuses on "
        "GBV/femicide/FGM. Eddy Kimani is a progressive mental health/fatherhood speaker. "
        "Amerix is a regressive masculinity/health influencer."
    )

    return f"""{CONTENT_SYSTEM}

{country_note}

{THEME_TAXONOMY}

=== CONTENT TO CODE ===
Content ID: {content_id}
Country: {country}
Context / Platform info: {context or '(not provided)'}
Content Text:
{text}

=== CODING QUESTIONS ===

Q1. Is this an attention-getter? (hooks/openers that capture audience attention before main content)
  [single]: Yes | No

Q1a. Attention-getting strategies used (ONLY if Q1=Yes; else [])
  [multi]: Compelling question | Use of all CAPS | Humor or sarcasm | Shares something violent or gross | Shares something sexual | Shares something surprising | Uses a news headline or social media trend as opener | Interesting visual or meme | Other

Q1b. If Q1a includes Other, describe strategy (8+ words). Else "".

Q2. Primary topic(s) of the content
  [multi]: Dating/marriage | Friends/socializing | Family/children | Money/status | Fitness/self-improvement | Mental health | Gender issues, e.g. equality | Social issues, e.g. corruption | Religion/morality | Gaming/technology | Other

Q2a. If Q2 includes Other, describe topic. Else "".

Q3. Type of content
  [single]: Interview/conversational content | Motivational/self-help content | Commentary/reaction content | Other

Q3a. If Q3=Other, describe content type. Else "".

Q4. Does the content address masculinity or gender norms?
  [single]: Yes, explicitly | Yes, implicitly | No

Q5. Type of masculinity/gender norms expressed
  [single]: More regressive/traditional/restrictive | More progressive/equitable/expansive | Mixed/unclear | Does not address masculinity or gender norms
  (If Q4=No → "Does not address masculinity or gender norms")

Q6. Does the content address what men should do?
  [single]: Yes | No | Unclear | Not applicable
  (If Q4=No → "Not applicable")

Q7. What men are told they do or need to do (ONLY if Q6=Yes; else ["Not applicable"])
  [multi]: Men need to dominate/lead | Men need to provide/succeed | Men are disadvantaged/victims | Men need to improve themselves | Men need to be fully self-reliant | Men need to be emotionally open | Men need to not show emotions | Men need to be equal partners | Mixed/unclear | Other | Not applicable

Q7a. If Q7 includes Other, describe directive (8+ words). Else "".

Q8. Problem identified in the content (select all that apply)
  [multi]: Kenyan or Nigerian political/social problems | Global political/social/cultural problems | Western political/social influence | Women/feminism | Men's behavior | Economic/status pressure | Mental health/emotional struggle | No clear problem is identified | Other

Q8a. If Q8 includes Other, describe problem. Else "".

Q9. Solution proposed in the content (select all that apply)
  [multi]: Social or political change | Assert dominance/control | More wealth/status | More self-discipline/fitness | More emotional growth/healing | More equality/respect for men | More equality/respect for women | Building community | No clear solution | Other

Q9a. If Q9 includes Other, describe solution. Else "".

Q10. Communication mode used (select all that apply)
  [multi]: Advice/instruction | Personal story | Commentary/opinion | Debate/argument | Humor/satire | Motivational speech | News/telling facts | Other

Q10a. If Q10 includes Other, describe mode. Else "".

Q11. Audience needs this content serves (select all that apply)
  [multi]: Entertainment/escapism | Information seeking | Connection/social interaction | Self expression/identity construction | Status seeking | Documentation of events | None of these apply

Q12. How claims are supported
  [single]: Generalizations about men/women | Personal experience | Stories about men/women | Cultural/social observations | Facts/statistics | Moral/religious claims | Mixed | No support | Other

Q12a. If Q12=Other, describe claim support. Else "".

Q13. How claims are justified
  [single]: No justification | Anecdotal examples | Presented as common sense | References data | References religion/tradition | References external sources, such as other influencers | Other

Q13a. If Q13=Other, describe justification. Else "".

Q14. Sentiment toward men expressed
  [single]: Negative | Positive | Mixed | Neutral | Unclear | Not mentioned

Q15. Sentiment toward women expressed
  [single]: Negative | Positive | Mixed | Neutral | Unclear | Not mentioned

Q16. Sentiment toward traditional gender norms expressed
  [single]: Negative | Positive | Mixed | Neutral | Unclear | Not mentioned

Q17. Is fear or threat used as a rhetorical device? (threatening what happens if men/women don't comply)
  [single]: Yes | Somewhat | No

Q18. Are calls to action present?
  [single]: Yes | No

Q18a. Types of calls to action (ONLY if Q18=Yes; else [])
  [multi]: Calls for audience to like the content | Calls for audience to share the content | Calls for audience to follow the speaker on social media | Calls for men to follow more traditional gender norms | Calls for men to follow more equitable gender norms | Calls for women to follow more traditional gender norms | Calls for women to follow more equitable gender norms | Calls for politicians or social figures to do something | Calls for audience to vote in a different way | Other

Q18b. If Q18a includes Other, describe call to action. Else "".

=== JSON RESPONSE ===
Return ONLY this JSON structure (exact keys, no extra keys, no markdown):
{{"q1":"","q1a":[],"q1b":"","q2":[],"q2a":"","q3":"","q3a":"","q4":"","q5":"","q6":"","q7":[],"q7a":"","q8":[],"q8a":"","q9":[],"q9a":"","q10":[],"q10a":"","q11":[],"q12":"","q12a":"","q13":"","q13a":"","q14":"","q15":"","q16":"","q17":"","q18":"","q18a":[],"q18b":""}}"""


def enforce_content_conditionals(r: dict) -> dict:
    v = CONTENT_VALID
    r["q1"]   = clean_single_field(r.get("q1", ""),   v["q1"],   "No")
    r["q1a"]  = clean_list_field(r.get("q1a", []),    v["q1a"])
    r["q1b"]  = clean_open_text(r.get("q1b", ""))
    r["q2"]   = clean_list_field(r.get("q2", []),     v["q2"])
    r["q2a"]  = clean_open_text(r.get("q2a", ""))
    r["q3"]   = clean_single_field(r.get("q3", ""),   v["q3"],   "Commentary/reaction content")
    r["q3a"]  = clean_open_text(r.get("q3a", ""))
    r["q4"]   = clean_single_field(r.get("q4", ""),   v["q4"],   "No")
    r["q5"]   = clean_single_field(r.get("q5", ""),   v["q5"],   "Does not address masculinity or gender norms")
    r["q6"]   = clean_single_field(r.get("q6", ""),   v["q6"],   "Not applicable")
    r["q7"]   = clean_list_field(r.get("q7", []),     v["q7"])
    r["q7a"]  = clean_open_text(r.get("q7a", ""))
    r["q8"]   = clean_list_field(r.get("q8", []),     v["q8"])
    r["q8a"]  = clean_open_text(r.get("q8a", ""))
    r["q9"]   = clean_list_field(r.get("q9", []),     v["q9"])
    r["q9a"]  = clean_open_text(r.get("q9a", ""))
    r["q10"]  = clean_list_field(r.get("q10", []),    v["q10"])
    r["q10a"] = clean_open_text(r.get("q10a", ""))
    r["q11"]  = clean_list_field(r.get("q11", []),    v["q11"])
    r["q12"]  = clean_single_field(r.get("q12", ""),  v["q12"],  "No support")
    r["q12a"] = clean_open_text(r.get("q12a", ""))
    r["q13"]  = clean_single_field(r.get("q13", ""),  v["q13"],  "No justification")
    r["q13a"] = clean_open_text(r.get("q13a", ""))
    r["q14"]  = clean_single_field(r.get("q14", ""),  v["q14"],  "Unclear")
    r["q15"]  = clean_single_field(r.get("q15", ""),  v["q15"],  "Unclear")
    r["q16"]  = clean_single_field(r.get("q16", ""),  v["q16"],  "Unclear")
    r["q17"]  = clean_single_field(r.get("q17", ""),  v["q17"],  "No")
    r["q18"]  = clean_single_field(r.get("q18", ""),  v["q18"],  "No")
    r["q18a"] = clean_list_field(r.get("q18a", []),   v["q18a"])
    r["q18b"] = clean_open_text(r.get("q18b", ""))

    # Q1 conditionals
    if r["q1"] == "No":
        r["q1a"] = []; r["q1b"] = ""
    else:
        if not r["q1a"]:
            r["q1a"] = ["Shares something surprising"]
        if "Other" not in r["q1a"]:
            r["q1b"] = ""
        else:
            r["q1b"] = ensure_substantive(r["q1b"], "Other attention-getting strategy not covered by listed options.")

    if "Other" not in r["q2"]:
        r["q2a"] = ""
    else:
        r["q2a"] = ensure_substantive(r["q2a"], "Other topic not covered by listed options in the codebook.")

    if r["q3"] != "Other":
        r["q3a"] = ""
    else:
        r["q3a"] = ensure_substantive(r["q3a"], "Other content type not covered by the listed options.")

    # Q4 → Q5, Q6, Q7
    if r["q4"] == "No":
        r["q5"] = "Does not address masculinity or gender norms"
        r["q6"] = "Not applicable"
        r["q7"] = ["Not applicable"]
        r["q7a"] = ""
    else:
        if r["q6"] != "Yes":
            r["q7"] = ["Not applicable"]
            r["q7a"] = ""
        else:
            r["q7"] = [x for x in r["q7"] if x != "Not applicable"] or ["Mixed/unclear"]
            if "Other" not in r["q7"]:
                r["q7a"] = ""
            else:
                r["q7a"] = ensure_substantive(r["q7a"], "Other directive for men not covered by listed options.")

    if "Other" not in r["q8"]:
        r["q8a"] = ""
    else:
        r["q8a"] = ensure_substantive(r["q8a"], "Other problem identified not covered by listed options.")
    if not r["q8"]:
        r["q8"] = ["No clear problem is identified"]

    if "Other" not in r["q9"]:
        r["q9a"] = ""
    else:
        r["q9a"] = ensure_substantive(r["q9a"], "Other solution proposed not covered by listed options.")
    if not r["q9"]:
        r["q9"] = ["No clear solution"]

    if "Other" not in r["q10"]:
        r["q10a"] = ""
    else:
        r["q10a"] = ensure_substantive(r["q10a"], "Other communication mode not covered by listed options.")
    if not r["q10"]:
        r["q10"] = ["Commentary/opinion"]

    if not r["q11"]:
        r["q11"] = ["None of these apply"]

    if r["q12"] != "Other":
        r["q12a"] = ""
    else:
        r["q12a"] = ensure_substantive(r["q12a"], "Other claim support mechanism not covered by listed options.")

    if r["q13"] != "Other":
        r["q13a"] = ""
    else:
        r["q13a"] = ensure_substantive(r["q13a"], "Other justification not covered by listed options.")

    # Q18 conditionals
    if r["q18"] == "No":
        r["q18a"] = []; r["q18b"] = ""
    else:
        if not r["q18a"]:
            r["q18a"] = ["Calls for audience to follow the speaker on social media"]
        if "Other" not in r["q18a"]:
            r["q18b"] = ""
        else:
            r["q18b"] = ensure_substantive(r["q18b"], "Other call to action not covered by listed options.")

    return r


def content_row_to_output(content_id: str, context: str, text: str, r: dict) -> dict:
    j = lambda lst: ", ".join(lst) if lst else ""
    return {
        "Content ID":                                content_id,
        "Context":                                   context,
        "Content Text / Description":                text,
        "Q1. Attention-getter (Yes/No)":             r["q1"],
        "Q1a. Attention-getting strategies":         j(r["q1a"]),
        "Q1b. Other strategy":                       r["q1b"],
        "Q2. Primary topic(s)":                      j(r["q2"]),
        "Q2a. Other topic":                          r["q2a"],
        "Q3. Type of content":                       r["q3"],
        "Q3a. Other content type":                   r["q3a"],
        "Q4. Addresses masculinity / gender norms":  r["q4"],
        "Q5. Type of masculinity / gender norms":    r["q5"],
        "Q6. Addresses what men should do":          r["q6"],
        "Q7. What men do or need to do":             j(r["q7"]),
        "Q7a. Other directive":                      r["q7a"],
        "Q8. Problem identified":                    j(r["q8"]),
        "Q8a. Other problem":                        r["q8a"],
        "Q9. Solution proposed":                     j(r["q9"]),
        "Q9a. Other solution":                       r["q9a"],
        "Q10. Communication mode":                   j(r["q10"]),
        "Q10a. Other communication mode":            r["q10a"],
        "Q11. Audience needs":                       j(r["q11"]),
        "Q12. How claims are supported":             r["q12"],
        "Q12a. Other claim support":                 r["q12a"],
        "Q13. How claims are justified":             r["q13"],
        "Q13a. Other justification":                 r["q13a"],
        "Q14. Sentiment toward men":                 r["q14"],
        "Q15. Sentiment toward women":               r["q15"],
        "Q16. Sentiment toward traditional gender norms": r["q16"],
        "Q17. Fear or threat used":                  r["q17"],
        "Q18. Calls to action present":              r["q18"],
        "Q18a. Types of calls to action":            j(r["q18a"]),
        "Q18b. Other call to action":                r["q18b"],
    }


async def run_content_analysis():
    print("\n" + "=" * 60)
    print("CONTENT ANALYSIS CODING")
    print("=" * 60)

    cache_path = CACHE_DIR / "llm_content_cache_claude.parquet"
    cache_df = load_cache(cache_path)
    cached_keys = set(cache_df["_cache_key"].tolist()) if "_cache_key" in cache_df.columns else set()
    new_cache_rows = []

    nigeria_df = pd.read_excel(CONTENT_MASTER, sheet_name="Nigeria — All Items", dtype=str).fillna("")
    kenya_df   = pd.read_excel(CONTENT_MASTER, sheet_name="Kenya — All Items",   dtype=str).fillna("")
    assert len(nigeria_df) == 200 and len(kenya_df) == 200

    sem = asyncio.Semaphore(CONCURRENCY)
    country_results = {}

    for country, df in [("Nigeria", nigeria_df), ("Kenya", kenya_df)]:
        print(f"\n--- {country} Content ({len(df)} rows) ---")
        results = {}

        # Pre-fill cached
        for _, row in df.iterrows():
            cid = str(row["Content ID"])
            key = cache_key(cid, country, "content")
            if key in cached_keys:
                hit = cache_df[cache_df["_cache_key"] == key]
                if len(hit):
                    results[cid] = json.loads(hit.iloc[0]["_result_json"])

        to_code = [row for _, row in df.iterrows()
                   if str(row["Content ID"]) not in results]
        print(f"  Cached: {len(results)}, To code: {len(to_code)}")

        async def code_one_content(row):
            async with sem:
                cid = str(row["Content ID"])
                ctx = str(row["Context"])
                txt = str(row["Content Text / Description"])
                prompt = build_content_prompt(cid, country, ctx, txt)
                raw = await call_with_retry(prompt, cid)
                coded = enforce_content_conditionals(raw)
                return cid, coded

        tasks = [code_one_content(row) for row in to_code]
        done = 0
        for coro in asyncio.as_completed(tasks):
            cid, coded = await coro
            results[cid] = coded
            new_cache_rows.append({
                "_cache_key":   cache_key(cid, country, "content"),
                "_country":     country,
                "_analysis":    "content",
                "_result_json": json.dumps(coded),
            })
            done += 1
            if done % 10 == 0 or done == len(to_code):
                print(f"  [{country}] {done}/{len(to_code)} coded")
                # Periodic cache save
                _save_new_cache(cache_df, new_cache_rows, cache_path)

        # Build ordered output
        ordered = []
        for _, row in df.iterrows():
            cid = str(row["Content ID"])
            ordered.append(content_row_to_output(
                cid, str(row["Context"]), str(row["Content Text / Description"]), results[cid]
            ))
        country_results[country] = pd.DataFrame(ordered, columns=CONTENT_OUTPUT_COLUMNS)
        assert len(country_results[country]) == 200

    # Save cache final
    _save_new_cache(cache_df, new_cache_rows, cache_path)

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meth = _content_methodology(run_date)

    for country in ["Nigeria", "Kenya"]:
        path = LLM_DIR / f"LLM Coding - Content Analysis - {country}.xlsx"
        write_excel_file(
            path,
            {f"{country} - LLM Coding": country_results[country]},
            run_date, meth
        )
        print(f"  {country}: {len(country_results[country])} rows ✓")

    combined_path = LLM_DIR / "LLM Coding - Content Analysis.xlsx"
    write_excel_file(
        combined_path,
        {"Nigeria - LLM Coding": country_results["Nigeria"],
         "Kenya - LLM Coding":   country_results["Kenya"]},
        run_date, meth
    )
    print(f"\nContent analysis complete. Files saved to {LLM_DIR}")


def _save_new_cache(old_cache: pd.DataFrame, new_rows: list, path: Path):
    if not new_rows:
        return
    new_df = pd.DataFrame(new_rows)
    combined = pd.concat([old_cache, new_df], ignore_index=True)
    combined = combined.drop_duplicates(subset=["_cache_key"], keep="last")
    save_cache(combined, path)


def _content_methodology(run_date: str) -> list:
    return [
        ["LLM Content Analysis — Methodology"],
        ["Field", "Value"],
        ["Model", "Claude Code (claude-sonnet-4-6 via Claude Desktop OAuth)"],
        ["Model tag", MODEL_TAG],
        ["Temperature", "Default (via CLI)"],
        ["Concurrency", str(CONCURRENCY)],
        ["Date run", run_date],
        ["Nigeria source rows", "200 (Nigeria — All Items sheet)"],
        ["Kenya source rows", "200 (Kenya — All Items sheet)"],
        ["Source", str(CONTENT_MASTER)],
        [],
        ["Notes"],
        ["LLM received only Content ID, Context, and Content Text. Human answers were not provided."],
        ["Conditional logic enforced in post-processing. Multi-select stored as comma-joined strings."],
        ["Results cached in parquet. Cache key includes model tag — change MODEL_TAG to bust cache."],
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  AUDIENCE ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

AUDIENCE_OUTPUT_COLUMNS = [
    "Comment ID", "Commenter Post URL", "Influencer's OG Post URL", "Comment Text",
    "Q1. Overall sentiment of comment",
    "Q2. Primary emotional tone",
    "Q3. Commenter's emotional response to content",
    "Q4. Mentions men, women, or gender norms",
    "Q5. Sentiment toward men / masculinity",
    "Q6. Sentiment toward women / femininity",
    "Q7. Main topic of comment", "Q7a. Other topic",
    "Q8. Commenter's stance toward the content",
    "Q9. If supporting, explain why",
    "Q10. Need the content is serving",
    "Q11. If challenging, explain why",
    "Q12. References personal experience",
    "Q13. Includes sexist / derogatory language",
    "Q14. Acquired new knowledge", "Q14a. If yes, what did they learn",
    "Q15. Changed attitudes", "Q15a. If yes, how did attitude change",
    "Q16. Opinion reinforced by content", "Q16a. If yes, what opinion was reinforced",
    "Q17. Calls to action present", "Q17a. If yes, what action is urged",
    "Q18. Shares information (fact, link, etc.)", "Q18a. If yes, what information",
    "Q19. Advocates for something", "Q19a. If yes, what do they advocate for",
    "Q20. Corrects content or other comments", "Q20a. If yes, what is incorrect / correction",
    "Q21. Commenter self-identifies",
    "Q21a. Profession mentioned", "Q21b. If yes, what profession",
    "Q21c. Location mentioned", "Q21d. If yes, what location",
    "Q21e. Race / ethnicity mentioned", "Q21f. If yes, what race / ethnicity",
    "Q21g. Gender mentioned", "Q21h. If yes, what gender",
]

AUDIENCE_VALID = {
    "q1":  {"Positive", "Negative", "Neutral", "Unclear"},
    "q2":  {"Joy", "Happiness", "Surprise", "Anger", "Fear", "Contempt",
            "Sadness", "Hope", "Empathy", "None of these"},
    "q3":  {"Feeling seen/understood", "Feeling unseen/misunderstood",
            "Feeling attacked", "Feeling objectified", "None of these"},
    "q4":  {"Yes", "No"},
    "q5":  {"Positive", "Negative", "Neutral", "Unclear", "Does not mention men/masculinity"},
    "q6":  {"Positive", "Negative", "Neutral", "Unclear", "Does not mention women/femininity"},
    "q7":  {"Dating/relationships/marriage", "Gender roles/norms", "Mental health/emotions",
            "Money/status", "Politics / social issues",
            "The speaker/creator of the content/influencer/the content",
            "Media/video games", "Other"},
    "q8":  {"Supporting", "Challenging", "Neutral", "Unclear"},
    "q10": {"Entertainment/escapism", "Information seeking", "Connection/social interaction",
            "Self expression/identity construction", "Status seeking",
            "Documentation of events", "None of these apply"},
    "q12": {"Yes", "No"}, "q13": {"Yes", "No"}, "q14": {"Yes", "No"},
    "q15": {"Yes", "No"}, "q16": {"Yes", "No"}, "q17": {"Yes", "No"},
    "q18": {"Yes", "No"}, "q19": {"Yes", "No"}, "q20": {"Yes", "No"},
    "q21": {"Yes", "No"}, "q21a": {"Yes", "No"}, "q21c": {"Yes", "No"},
    "q21e": {"Yes", "No"}, "q21g": {"Yes", "No"},
}

AUDIENCE_SYSTEM = """You are a precise qualitative research coder for a study on how audiences respond to masculinity-focused social media content in Nigeria and Kenya. Code each comment using the structured codebook below.

Critical rules:
- Answer ONLY using the exact options listed. Do not invent options.
- Multi-select fields: return JSON array. Single-select: return string. Open-text: return string ("" if not applicable).
- Return ONLY valid JSON — no markdown, no commentary.
- Open-text explanations (Q9, Q11, Q14a–Q20a) must be substantive (at least 1 clear sentence, 8+ words). Never write vague one-word answers like "they agree" or "they support".
- Q8 STANCE RULES (very important):
  * Supporting = commenter agrees with, validates, or amplifies the influencer's message
  * Challenging = commenter disagrees, pushes back, corrects, or offers a counter-perspective
  * Neutral = commenter engages without taking a stance (asks questions, describes, shares info)
  * Unclear = insufficient content to determine stance
  * NOTE: A commenter may share personal experience WHILE supporting OR challenging — code based on their stance toward the content, not just what they say about themselves
- Q21 SELF-IDENTIFICATION RULES (very important):
  * Code Q21=Yes when the commenter reveals ANY personal info about themselves, including:
    - Explicit gender statements: "I am a man/woman/father/mother"
    - RELATIONAL terms that imply gender: "my husband", "my wife", "my boyfriend", "my girlfriend", "I am married to [he/she]", "my late husband", "my ex-wife"
    - Profession: "as a doctor", "I am a teacher", "I work as..."
    - Location: "here in Lagos", "I'm from Kenya", "living in Nairobi"
    - Marital status: "I am married", "I am divorced", "as a widow"
    - Family role: "as a father", "as a mother", "as a parent"
  * For Q21g (gender mentioned): code Yes if commenter uses ANY relational term that implies gender (e.g., "my husband" implies the commenter is likely a woman, "my wife" implies a man). Code the implied gender in Q21h.
  * Do NOT code Q21=No just because the commenter didn't explicitly say "I am a man/woman" — relational language counts."""


def build_audience_prompt(comment_id: str, country: str,
                           commenter_url: str, og_url: str, comment_text: str) -> str:
    country_note = (
        "Nigeria context: Pidgin/Yoruba/Igbo/Hausa terms may appear. Faith framing is common. "
        "Creators: Agba John Doe, Banky Wellington, Deyemi Okanlawon, Ebuka Obi-Uchendu, Shola, Wizarab."
        if country == "Nigeria" else
        "Kenya context: Sheng/Swahili terms may appear. Andrew Kibe/Jagero is a regressive manosphere influencer. "
        "Rixpoet/Onyango is progressive mental health. Philip Karanja focuses on GBV/femicide. "
        "Eddy Kimani is progressive mental health/fatherhood. Amerix is a regressive masculinity influencer."
    )

    # Extract creator hint from URL
    og_lower = og_url.lower()
    creator_hint = ""
    hints = {
        "bankywellington": "Banky Wellington (Nigerian musician/politician, progressive)",
        "deyemi": "Deyemi Okanlawon (Nigerian actor, progressive)",
        "ebuka": "Ebuka Obi-Uchendu (Nigerian TV presenter)",
        "wizarab": "Wizarab (Nigerian influencer)",
        "andrewkibe": "Andrew Kibe (Kenyan manosphere, regressive)",
        "jagero": "Andrew Kibe/Jagero (Kenyan manosphere, regressive)",
        "rixpoet": "Rixpoet (Kenyan progressive mental health)",
        "onyango": "Onyango Otieno (Kenyan progressive mental health)",
        "philipkaranja": "Philip Karanja (Kenyan GBV/femicide advocate)",
        "eddykimani": "Eddy Kimani (Kenyan progressive mental health/fatherhood)",
        "amerix": "Amerix (Kenyan regressive masculinity influencer)",
    }
    for kw, hint in hints.items():
        if kw in og_lower:
            creator_hint = f"\nContent creator: {hint}"
            break

    return f"""{AUDIENCE_SYSTEM}

{country_note}

=== COMMENT TO CODE ===
Comment ID: {comment_id}
Country: {country}
Commenter URL: {commenter_url or '(not provided)'}{creator_hint}
OG Post URL: {og_url or '(not provided)'}
Comment Text:
{comment_text}

=== CODING QUESTIONS ===

Q1. Overall sentiment of the comment
  [single]: Positive | Negative | Neutral | Unclear

Q2. Primary emotional tone of the comment
  [single]: Joy | Happiness | Surprise | Anger | Fear | Contempt | Sadness | Hope | Empathy | None of these

Q3. Commenter's emotional response to the content (may select multiple)
  [multi]: Feeling seen/understood | Feeling unseen/misunderstood | Feeling attacked | Feeling objectified | None of these

Q4. Does the comment mention men, women, or gender norms?
  [single]: Yes | No

Q5. Sentiment toward men / masculinity
  [single]: Positive | Negative | Neutral | Unclear | Does not mention men/masculinity

Q6. Sentiment toward women / femininity
  [single]: Positive | Negative | Neutral | Unclear | Does not mention women/femininity

Q7. Main topic(s) of the comment (select all that apply)
  [multi]: Dating/relationships/marriage | Gender roles/norms | Mental health/emotions | Money/status | Politics / social issues | The speaker/creator of the content/influencer/the content | Media/video games | Other

Q7a. If Q7 includes Other, describe topic (8+ words). Else "".

Q8. Commenter's stance toward the content they are responding to
  [single]: Supporting | Challenging | Neutral | Unclear
  (See stance rules above — read the full comment before deciding)

Q9. If Q8=Supporting: explain in 1-2 specific sentences WHY the commenter supports the content (their reasoning, not just 'they agree'). Else "".

Q10. Need the comment/content is serving for the commenter (select all that apply)
  [multi]: Entertainment/escapism | Information seeking | Connection/social interaction | Self expression/identity construction | Status seeking | Documentation of events | None of these apply

Q11. If Q8=Challenging: explain in 1-2 specific sentences WHY the commenter challenges the content (their objection, not just 'they disagree'). Else "".

Q12. Does the comment reference personal experience?
  [single]: Yes | No

Q13. Does the comment include sexist or derogatory language?
  [single]: Yes | No

Q14. Does the comment suggest the commenter acquired new knowledge from the content?
  [single]: Yes | No

Q14a. If Q14=Yes: what did they learn? (1-2 sentences). Else "".

Q15. Does the comment suggest the commenter's attitude changed because of the content?
  [single]: Yes | No

Q15a. If Q15=Yes: how did their attitude change? (1-2 sentences). Else "".

Q16. Does the comment suggest an existing opinion was reinforced by the content?
  [single]: Yes | No

Q16a. If Q16=Yes: what opinion was reinforced? (1-2 sentences). Else "".

Q17. Are there calls to action in the comment?
  [single]: Yes | No

Q17a. If Q17=Yes: what action is being urged? (1-2 sentences). Else "".

Q18. Does the comment share information (a fact, link, statistic, or external reference)?
  [single]: Yes | No

Q18a. If Q18=Yes: what information is shared? (1-2 sentences). Else "".

Q19. Does the commenter advocate for something (a cause, position, or change)?
  [single]: Yes | No

Q19a. If Q19=Yes: what do they advocate for? (1-2 sentences). Else "".

Q20. Does the comment correct the content or other comments?
  [single]: Yes | No

Q20a. If Q20=Yes: what is incorrect and what is the correction? (1-2 sentences). Else "".

Q21. Does the commenter self-identify in any way? (See Q21 rules above — relational terms count!)
  [single]: Yes | No

Q21a. If Q21=Yes: does the commenter mention their profession? [single]: Yes | No. (If Q21=No → "No")

Q21b. If Q21a=Yes: what profession? Else "".

Q21c. If Q21=Yes: does the commenter mention their location? [single]: Yes | No. (If Q21=No → "No")

Q21d. If Q21c=Yes: what location? Else "".

Q21e. If Q21=Yes: does the commenter mention their race or ethnicity? [single]: Yes | No. (If Q21=No → "No")

Q21f. If Q21e=Yes: what race/ethnicity? Else "".

Q21g. If Q21=Yes: does the commenter mention their gender (explicitly OR via relational terms)? [single]: Yes | No. (If Q21=No → "No")

Q21h. If Q21g=Yes: what gender? (Use "man" or "woman" — infer from relational terms if needed). Else "".

=== JSON RESPONSE ===
Return ONLY this JSON (exact keys, no markdown):
{{"q1":"","q2":"","q3":[],"q4":"","q5":"","q6":"","q7":[],"q7a":"","q8":"","q9":"","q10":[],"q11":"","q12":"","q13":"","q14":"","q14a":"","q15":"","q15a":"","q16":"","q16a":"","q17":"","q17a":"","q18":"","q18a":"","q19":"","q19a":"","q20":"","q20a":"","q21":"","q21a":"","q21b":"","q21c":"","q21d":"","q21e":"","q21f":"","q21g":"","q21h":""}}"""


def enforce_audience_conditionals(r: dict) -> dict:
    v = AUDIENCE_VALID
    r["q1"]   = clean_single_field(r.get("q1",  ""), v["q1"],   "Unclear")
    r["q2"]   = clean_single_field(r.get("q2",  ""), v["q2"],   "None of these")
    r["q3"]   = clean_list_field(r.get("q3",   []), v["q3"])
    r["q4"]   = clean_single_field(r.get("q4",  ""), v["q4"],   "No")
    r["q5"]   = clean_single_field(r.get("q5",  ""), v["q5"],   "Does not mention men/masculinity")
    r["q6"]   = clean_single_field(r.get("q6",  ""), v["q6"],   "Does not mention women/femininity")
    r["q7"]   = clean_list_field(r.get("q7",   []), v["q7"])
    r["q7a"]  = clean_open_text(r.get("q7a",   ""))
    r["q8"]   = clean_single_field(r.get("q8",  ""), v["q8"],   "Unclear")
    r["q9"]   = clean_open_text(r.get("q9",    ""))
    r["q10"]  = clean_list_field(r.get("q10",  []), v["q10"])
    r["q11"]  = clean_open_text(r.get("q11",   ""))
    r["q12"]  = clean_single_field(r.get("q12", ""), v["q12"],  "No")
    r["q13"]  = clean_single_field(r.get("q13", ""), v["q13"],  "No")
    r["q14"]  = clean_single_field(r.get("q14", ""), v["q14"],  "No")
    r["q14a"] = clean_open_text(r.get("q14a",  ""))
    r["q15"]  = clean_single_field(r.get("q15", ""), v["q15"],  "No")
    r["q15a"] = clean_open_text(r.get("q15a",  ""))
    r["q16"]  = clean_single_field(r.get("q16", ""), v["q16"],  "No")
    r["q16a"] = clean_open_text(r.get("q16a",  ""))
    r["q17"]  = clean_single_field(r.get("q17", ""), v["q17"],  "No")
    r["q17a"] = clean_open_text(r.get("q17a",  ""))
    r["q18"]  = clean_single_field(r.get("q18", ""), v["q18"],  "No")
    r["q18a"] = clean_open_text(r.get("q18a",  ""))
    r["q19"]  = clean_single_field(r.get("q19", ""), v["q19"],  "No")
    r["q19a"] = clean_open_text(r.get("q19a",  ""))
    r["q20"]  = clean_single_field(r.get("q20", ""), v["q20"],  "No")
    r["q20a"] = clean_open_text(r.get("q20a",  ""))
    r["q21"]  = clean_single_field(r.get("q21", ""), v["q21"],  "No")
    r["q21a"] = clean_single_field(r.get("q21a",""), v["q21a"], "No")
    r["q21b"] = clean_open_text(r.get("q21b",  ""))
    r["q21c"] = clean_single_field(r.get("q21c",""), v["q21c"], "No")
    r["q21d"] = clean_open_text(r.get("q21d",  ""))
    r["q21e"] = clean_single_field(r.get("q21e",""), v["q21e"], "No")
    r["q21f"] = clean_open_text(r.get("q21f",  ""))
    r["q21g"] = clean_single_field(r.get("q21g",""), v["q21g"], "No")
    r["q21h"] = clean_open_text(r.get("q21h",  ""))

    if not r["q3"]:
        r["q3"] = ["None of these"]
    if not r["q7"]:
        r["q7"] = ["The speaker/creator of the content/influencer/the content"]
    if "Other" not in r["q7"]:
        r["q7a"] = ""
    else:
        r["q7a"] = ensure_substantive(r["q7a"], "Other topic not covered by listed options.")

    if r["q8"] != "Supporting":
        r["q9"] = ""
    else:
        r["q9"] = ensure_substantive(r["q9"], "The commenter supports the content and validates its core message or reasoning.")

    if r["q8"] != "Challenging":
        r["q11"] = ""
    else:
        r["q11"] = ensure_substantive(r["q11"], "The commenter challenges the content by offering a counter-perspective or correction.")

    if not r["q10"]:
        r["q10"] = ["None of these apply"]

    for yesno_key, text_key in [("q14","q14a"),("q15","q15a"),("q16","q16a"),
                                  ("q17","q17a"),("q18","q18a"),("q19","q19a"),("q20","q20a")]:
        if r[yesno_key] == "No":
            r[text_key] = ""

    if r["q21"] == "No":
        r["q21a"] = "No"; r["q21b"] = ""
        r["q21c"] = "No"; r["q21d"] = ""
        r["q21e"] = "No"; r["q21f"] = ""
        r["q21g"] = "No"; r["q21h"] = ""
    else:
        if r["q21a"] == "No":
            r["q21b"] = ""
        if r["q21c"] == "No":
            r["q21d"] = ""
        if r["q21e"] == "No":
            r["q21f"] = ""
        if r["q21g"] == "No":
            r["q21h"] = ""

    return r


def audience_row_to_output(comment_id: str, commenter_url: str,
                            og_url: str, comment_text: str, r: dict) -> dict:
    j = lambda lst: ", ".join(lst) if lst else ""
    return {
        "Comment ID":                                    comment_id,
        "Commenter Post URL":                            commenter_url,
        "Influencer's OG Post URL":                      og_url,
        "Comment Text":                                  comment_text,
        "Q1. Overall sentiment of comment":              r["q1"],
        "Q2. Primary emotional tone":                    r["q2"],
        "Q3. Commenter's emotional response to content": j(r["q3"]),
        "Q4. Mentions men, women, or gender norms":      r["q4"],
        "Q5. Sentiment toward men / masculinity":        r["q5"],
        "Q6. Sentiment toward women / femininity":       r["q6"],
        "Q7. Main topic of comment":                     j(r["q7"]),
        "Q7a. Other topic":                              r["q7a"],
        "Q8. Commenter's stance toward the content":     r["q8"],
        "Q9. If supporting, explain why":                r["q9"],
        "Q10. Need the content is serving":              j(r["q10"]),
        "Q11. If challenging, explain why":              r["q11"],
        "Q12. References personal experience":           r["q12"],
        "Q13. Includes sexist / derogatory language":    r["q13"],
        "Q14. Acquired new knowledge":                   r["q14"],
        "Q14a. If yes, what did they learn":             r["q14a"],
        "Q15. Changed attitudes":                        r["q15"],
        "Q15a. If yes, how did attitude change":         r["q15a"],
        "Q16. Opinion reinforced by content":            r["q16"],
        "Q16a. If yes, what opinion was reinforced":     r["q16a"],
        "Q17. Calls to action present":                  r["q17"],
        "Q17a. If yes, what action is urged":            r["q17a"],
        "Q18. Shares information (fact, link, etc.)":    r["q18"],
        "Q18a. If yes, what information":                r["q18a"],
        "Q19. Advocates for something":                  r["q19"],
        "Q19a. If yes, what do they advocate for":       r["q19a"],
        "Q20. Corrects content or other comments":       r["q20"],
        "Q20a. If yes, what is incorrect / correction":  r["q20a"],
        "Q21. Commenter self-identifies":                r["q21"],
        "Q21a. Profession mentioned":                    r["q21a"],
        "Q21b. If yes, what profession":                 r["q21b"],
        "Q21c. Location mentioned":                      r["q21c"],
        "Q21d. If yes, what location":                   r["q21d"],
        "Q21e. Race / ethnicity mentioned":              r["q21e"],
        "Q21f. If yes, what race / ethnicity":           r["q21f"],
        "Q21g. Gender mentioned":                        r["q21g"],
        "Q21h. If yes, what gender":                     r["q21h"],
    }


async def run_audience_analysis():
    print("\n" + "=" * 60)
    print("AUDIENCE ANALYSIS CODING")
    print("=" * 60)

    cache_path = CACHE_DIR / "llm_audience_cache_claude.parquet"
    cache_df = load_cache(cache_path)
    cached_keys = set(cache_df["_cache_key"].tolist()) if "_cache_key" in cache_df.columns else set()
    new_cache_rows = []

    nigeria_df = pd.read_excel(AUDIENCE_MASTER, sheet_name="Nigeria — All Items", dtype=str).fillna("")
    kenya_df   = pd.read_excel(AUDIENCE_MASTER, sheet_name="Kenya — All Items",   dtype=str).fillna("")
    assert len(nigeria_df) == 200 and len(kenya_df) == 200

    sem = asyncio.Semaphore(CONCURRENCY)
    country_results = {}

    for country, df in [("Nigeria", nigeria_df), ("Kenya", kenya_df)]:
        print(f"\n--- {country} Audience ({len(df)} rows) ---")
        results = {}

        for _, row in df.iterrows():
            cid = str(row["Comment ID"])
            key = cache_key(cid, country, "audience")
            if key in cached_keys:
                hit = cache_df[cache_df["_cache_key"] == key]
                if len(hit):
                    results[cid] = json.loads(hit.iloc[0]["_result_json"])

        to_code = [row for _, row in df.iterrows()
                   if str(row["Comment ID"]) not in results]
        print(f"  Cached: {len(results)}, To code: {len(to_code)}")

        async def code_one_audience(row):
            async with sem:
                cid   = str(row["Comment ID"])
                curl  = str(row["Commenter Post URL"])
                ogurl = str(row["Influencer's OG Post URL"])
                txt   = str(row["Comment Text"])
                prompt = build_audience_prompt(cid, country, curl, ogurl, txt)
                raw = await call_with_retry(prompt, cid)
                coded = enforce_audience_conditionals(raw)
                return cid, coded

        tasks = [code_one_audience(row) for row in to_code]
        done = 0
        for coro in asyncio.as_completed(tasks):
            cid, coded = await coro
            results[cid] = coded
            new_cache_rows.append({
                "_cache_key":   cache_key(cid, country, "audience"),
                "_country":     country,
                "_analysis":    "audience",
                "_result_json": json.dumps(coded),
            })
            done += 1
            if done % 10 == 0 or done == len(to_code):
                print(f"  [{country}] {done}/{len(to_code)} coded")
                _save_new_cache(cache_df, new_cache_rows, cache_path)

        ordered = []
        for _, row in df.iterrows():
            cid = str(row["Comment ID"])
            ordered.append(audience_row_to_output(
                cid, str(row["Commenter Post URL"]),
                str(row["Influencer's OG Post URL"]),
                str(row["Comment Text"]), results[cid]
            ))
        country_results[country] = pd.DataFrame(ordered, columns=AUDIENCE_OUTPUT_COLUMNS)
        assert len(country_results[country]) == 200

    _save_new_cache(cache_df, new_cache_rows, cache_path)

    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meth = _audience_methodology(run_date)

    for country in ["Nigeria", "Kenya"]:
        path = LLM_DIR / f"LLM Coding - Audience Analysis - {country}.xlsx"
        write_excel_file(
            path,
            {f"{country} - LLM Coding": country_results[country]},
            run_date, meth
        )
        print(f"  {country}: {len(country_results[country])} rows ✓")

    combined_path = LLM_DIR / "LLM Coding - Audience Analysis.xlsx"
    write_excel_file(
        combined_path,
        {"Nigeria - LLM Coding": country_results["Nigeria"],
         "Kenya - LLM Coding":   country_results["Kenya"]},
        run_date, meth
    )
    print(f"\nAudience analysis complete. Files saved to {LLM_DIR}")


def _audience_methodology(run_date: str) -> list:
    return [
        ["LLM Audience Analysis — Methodology"],
        ["Field", "Value"],
        ["Model", "Claude Code (claude-sonnet-4-6 via Claude Desktop OAuth)"],
        ["Model tag", MODEL_TAG],
        ["Temperature", "Default (via CLI)"],
        ["Concurrency", str(CONCURRENCY)],
        ["Date run", run_date],
        ["Nigeria source rows", "200 (Nigeria — All Items sheet)"],
        ["Kenya source rows", "200 (Kenya — All Items sheet)"],
        ["Source", str(AUDIENCE_MASTER)],
        [],
        ["Key improvements over GPT-4o-mini baseline"],
        ["Q21 self-ID: relational terms (my husband/wife, I am married) now coded as Q21=Yes, Q21g=Yes"],
        ["Q8 stance: explicit Supporting/Challenging rules provided with examples"],
        ["Q9/Q11: substantive explanation enforced (minimum 8 words)"],
        ["Conditional logic enforced in post-processing."],
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["content", "audience", "all"], default="all")
    args = parser.parse_args()

    t0 = time.time()
    print(f"Starting at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Mode: {args.mode} | Concurrency: {CONCURRENCY} | Cache tag: {MODEL_TAG}")

    if args.mode in ("content", "all"):
        await run_content_analysis()

    if args.mode in ("audience", "all"):
        await run_audience_analysis()

    elapsed = time.time() - t0
    print(f"\nAll done in {elapsed/60:.1f} min")


if __name__ == "__main__":
    asyncio.run(main())
