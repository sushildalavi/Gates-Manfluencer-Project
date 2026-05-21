from __future__ import annotations

from pathlib import Path
import re
import json
from collections import defaultdict, Counter
from dataclasses import dataclass
from typing import Dict, List, Tuple, Set

import pandas as pd
from openpyxl import load_workbook, Workbook

ROOT = Path('/Users/sushildalavi/Desktop/NLC/Gates-Manfluencer-Project')
HUMAN_DIR = ROOT / 'Codebooks' / 'Human Codebooks'
LLM_DIR = ROOT / 'Codebooks' / 'LLM Codebook'
OUT_DIR = ROOT / 'Codebooks' / 'Human Codebooks Cleaning Reports'
OUT_DIR.mkdir(parents=True, exist_ok=True)

Q_RE = re.compile(r'^(q\d+[a-z]?)\b', re.I)


def nt(s: str) -> str:
    s = str(s)
    s = s.replace('\u00a0', ' ').replace('\u200b', '')
    s = s.replace('\u2018', "'").replace('\u2019', "'")
    s = s.replace('\u201c', '"').replace('\u201d', '"')
    s = s.replace('\u2013', '-').replace('\u2014', '-')
    s = re.sub(r'\s+', ' ', s.strip())
    return s.strip(' ,;')


def nk(s: str) -> str:
    s = nt(s).lower().replace('&', 'and')
    s = re.sub(r'[^a-z0-9]+', '', s)
    return s


def is_multiselect(header: str) -> bool:
    h = header.lower()
    return ('choose all that apply' in h) or ('select all that apply' in h)


def split_multi(v: str, header: str) -> List[str]:
    s = nt(v)
    # Protect known comma-bearing canonical options
    if 'q2. what is/are the primary topic(s) of the content' in header.lower():
        s = s.replace('Gender issues, e.g. equality', 'Gender issues__EG__equality')
        s = s.replace('Social issues, e.g. corruption', 'Social issues__EG__corruption')
    parts = [nt(x) for x in re.split(r'[;,]', s) if nt(x)]
    out = []
    for p in parts:
        p = p.replace('Gender issues__EG__equality', 'Gender issues, e.g. equality')
        p = p.replace('Social issues__EG__corruption', 'Social issues, e.g. corruption')
        out.append(p)
    return out


def detect_header_row(ws) -> int | None:
    for r in range(1, min(30, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if sum(1 for v in vals if v not in (None, '')) >= 3:
            return r
    return None


def sheet_to_df(path: Path, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name]
    hr = detect_header_row(ws)
    if hr is None:
        return pd.DataFrame()
    headers = [nt(ws.cell(hr, c).value) if ws.cell(hr, c).value is not None else '' for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(hr + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v in (None, '') for v in vals):
            continue
        rows.append(vals)
    df = pd.DataFrame(rows, columns=headers)
    return df


def identify_id_col(df: pd.DataFrame) -> str:
    cands = [c for c in df.columns if 'comment id' in c.lower() or 'content id' in c.lower()]
    if not cands:
        raise ValueError('No ID column found')
    return cands[0]


def question_columns(df: pd.DataFrame) -> Dict[str, str]:
    out = {}
    for c in df.columns:
        m = Q_RE.match(c.lower())
        if m:
            q = m.group(1).lower()
            if q not in out:
                out[q] = c
    return out


@dataclass
class TrackConfig:
    name: str
    human_glob: str
    llm_file: Path
    llm_sheets: Tuple[str, str]


TRACKS = [
    TrackConfig(
        name='audience',
        human_glob='Audience Analysis Codebooks/*.xlsx',
        llm_file=LLM_DIR / 'LLM Coding - Audience Analysis.xlsx',
        llm_sheets=('Nigeria - LLM Coding', 'Kenya - LLM Coding'),
    ),
    TrackConfig(
        name='content',
        human_glob='Content Analysis Codebooks/*.xlsx',
        llm_file=LLM_DIR / 'LLM Coding - Content Analysis.xlsx',
        llm_sheets=('Nigeria - LLM Coding', 'Kenya - LLM Coding'),
    ),
]


def cohen_kappa(y_true: List[str], y_pred: List[str]) -> float | None:
    if len(y_true) == 0:
        return None
    labels = sorted(set(y_true) | set(y_pred))
    if len(labels) <= 1:
        return 1.0
    idx = {l: i for i, l in enumerate(labels)}
    n = len(labels)
    mat = [[0] * n for _ in range(n)]
    for t, p in zip(y_true, y_pred):
        mat[idx[t]][idx[p]] += 1
    N = sum(sum(r) for r in mat)
    po = sum(mat[i][i] for i in range(n)) / N
    row = [sum(mat[i][j] for j in range(n)) / N for i in range(n)]
    col = [sum(mat[i][j] for i in range(n)) / N for j in range(n)]
    pe = sum(row[i] * col[i] for i in range(n))
    if abs(1 - pe) < 1e-12:
        return None
    return (po - pe) / (1 - pe)


def agreement_metrics_single(rows: List[Tuple[str, str]]) -> Dict[str, float | int | None]:
    y_true = [t for t, _ in rows]
    y_pred = [p for _, p in rows]
    n = len(rows)
    exact = sum(1 for t, p in rows if nk(t) == nk(p))
    return {
        'n': n,
        'exact_match': exact,
        'accuracy': exact / n if n else None,
        'kappa': cohen_kappa([nk(x) for x in y_true], [nk(x) for x in y_pred]),
    }


def agreement_metrics_multi(rows: List[Tuple[Set[str], Set[str]]]) -> Dict[str, float | int | None]:
    n = len(rows)
    if n == 0:
        return {'n': 0, 'exact_set_match': 0, 'exact_rate': None, 'mean_jaccard': None, 'micro_precision': None, 'micro_recall': None, 'micro_f1': None}
    exact = 0
    jac = []
    tp = fp = fn = 0
    for tset, pset in rows:
        if tset == pset:
            exact += 1
        inter = len(tset & pset)
        union = len(tset | pset)
        jac.append(inter / union if union else 1.0)
        tp += inter
        fp += len(pset - tset)
        fn += len(tset - pset)
    prec = tp / (tp + fp) if (tp + fp) else None
    rec = tp / (tp + fn) if (tp + fn) else None
    f1 = (2 * prec * rec / (prec + rec)) if (prec is not None and rec is not None and (prec + rec) > 0) else None
    return {
        'n': n,
        'exact_set_match': exact,
        'exact_rate': exact / n,
        'mean_jaccard': sum(jac) / n,
        'micro_precision': prec,
        'micro_recall': rec,
        'micro_f1': f1,
    }


def build_master_and_agreement(track: TrackConfig):
    human_files = sorted((HUMAN_DIR / track.human_glob.split('/')[0]).glob('*.xlsx'))

    # country containers
    country_data = { 'Nigeria': [], 'Kenya': [] }

    for fp in human_files:
        coder = fp.stem.split(' - ')[0]
        for country in ('Nigeria', 'Kenya'):
            df = sheet_to_df(fp, country)
            if df.empty:
                continue
            id_col = identify_id_col(df)
            qcols = question_columns(df)
            df = df.copy()
            df[id_col] = df[id_col].astype(str).map(nt)
            df = df[df[id_col] != '']
            country_data[country].append((coder, fp, df, id_col, qcols))

    master_wb = Workbook()
    # remove default
    default = master_wb.active
    master_wb.remove(default)

    overall_records = []
    q_stats = []
    coder_stats = []

    llm_wb = load_workbook(track.llm_file, data_only=False)

    for country, llm_sheet in [('Nigeria', track.llm_sheets[0]), ('Kenya', track.llm_sheets[1])]:
        coder_items = country_data[country]
        if not coder_items:
            continue

        # unify id + question columns by first coder schema
        _, _, df0, id_col0, q0 = coder_items[0]

        # gather IDs per coder
        id_to_rows = defaultdict(list)
        for coder, fp, df, id_col, qcols in coder_items:
            for _, row in df.iterrows():
                _id = nt(str(row[id_col]))
                if not _id:
                    continue
                id_to_rows[_id].append((coder, fp, row, qcols, id_col))

        overlap_ids = {k for k, v in id_to_rows.items() if len(v) > 1}
        non_overlap_ids = [k for k, v in id_to_rows.items() if len(v) == 1]

        # build master rows (exclude overlap/intercoder)
        master_rows = []
        for _id in non_overlap_ids:
            coder, fp, row, qcols, id_col = id_to_rows[_id][0]
            rec = {col: row[col] if col in row else None for col in df0.columns}
            rec['Coder'] = coder
            rec['Source Workbook'] = fp.name
            master_rows.append(rec)

        master_df = pd.DataFrame(master_rows)
        if not master_df.empty:
            # write sheet
            ws = master_wb.create_sheet(f'{country}')
            cols = list(master_df.columns)
            ws.append(cols)
            for _, rr in master_df.iterrows():
                ws.append([rr.get(c, None) for c in cols])

        # load llm sheet
        llm_df = sheet_to_df(track.llm_file, llm_sheet)
        llm_id_col = identify_id_col(llm_df)
        llm_q = question_columns(llm_df)
        llm_df = llm_df.copy()
        llm_df[llm_id_col] = llm_df[llm_id_col].astype(str).map(nt)
        llm_df = llm_df[llm_df[llm_id_col] != '']
        llm_by_id = {nt(str(r[llm_id_col])): r for _, r in llm_df.iterrows()}

        # create consensus human rows for agreement:
        # - non-overlap: single coder row
        # - overlap: majority vote (single labels) / union-majority for multiselect
        consensus = {}
        qkeys = sorted(set().union(*[set(item[4].keys()) for item in coder_items]))

        for _id, items in id_to_rows.items():
            cons = {}
            for q in qkeys:
                vals = []
                hdr = None
                multi = False
                for coder, fp, row, qcols, id_col in items:
                    if q in qcols:
                        hdr = qcols[q]
                        v = row.get(hdr, None)
                        if isinstance(v, str) and nt(v):
                            vals.append(nt(v))
                            if is_multiselect(hdr):
                                multi = True
                if not vals:
                    continue
                if not multi:
                    cnt = Counter(vals)
                    cons[q] = cnt.most_common(1)[0][0]
                else:
                    # token majority across coder sets
                    token_counts = Counter()
                    n_coders = len(vals)
                    for v in vals:
                        toks = set(split_multi(v, hdr or q))
                        for t in toks:
                            token_counts[t] += 1
                    keep = [t for t, c in token_counts.items() if c >= max(1, (n_coders + 1) // 2)]
                    if not keep:
                        keep = [t for t, _ in token_counts.most_common(1)]
                    cons[q] = ', '.join(sorted(set(keep), key=lambda x: x.lower()))
            consensus[_id] = cons

        # agreement calculations: consensus human vs llm on shared ids
        shared_ids = sorted(set(consensus.keys()) & set(llm_by_id.keys()))
        for q in sorted(set(qkeys) & set(llm_q.keys())):
            human_vals_single = []
            human_vals_multi = []
            is_multi_q = False
            # infer multiselect from any coder header containing choose all
            for coder, fp, df, id_col, qcols in coder_items:
                if q in qcols and is_multiselect(qcols[q]):
                    is_multi_q = True
                    break

            for _id in shared_ids:
                hval = consensus.get(_id, {}).get(q)
                lrow = llm_by_id.get(_id)
                if hval is None or lrow is None:
                    continue
                lcol = llm_q[q]
                lval = lrow.get(lcol, None)
                if not isinstance(lval, str) or not nt(lval):
                    continue
                lval = nt(lval)
                if is_multi_q:
                    human_vals_multi.append((set(split_multi(hval, q)), set(split_multi(lval, q))))
                else:
                    human_vals_single.append((hval, lval))

            if is_multi_q:
                m = agreement_metrics_multi(human_vals_multi)
                mtype = 'multi'
            else:
                m = agreement_metrics_single(human_vals_single)
                mtype = 'single'
            q_stats.append({
                'track': track.name,
                'country': country,
                'question': q,
                'question_human_header': q0.get(q, ''),
                'question_llm_header': llm_q.get(q, ''),
                'type': mtype,
                **m,
                'shared_ids': len(shared_ids),
                'overlap_ids_intercoder': len(overlap_ids),
                'non_overlap_ids_master': len(non_overlap_ids),
            })

        # coder-level diagnostics vs llm on that coder rows only
        for coder, fp, df, id_col, qcols in coder_items:
            ids = [nt(str(x)) for x in df[id_col].dropna().astype(str).tolist() if nt(str(x))]
            ids = sorted(set(ids) & set(llm_by_id.keys()))
            for q in sorted(set(qcols.keys()) & set(llm_q.keys())):
                hcol = qcols[q]
                lcol = llm_q[q]
                is_multi_q = is_multiselect(hcol)
                rows_single, rows_multi = [], []
                # index human by id
                idx = {}
                for _, rr in df.iterrows():
                    _id = nt(str(rr[id_col]))
                    if _id:
                        idx[_id] = rr
                for _id in ids:
                    rr = idx.get(_id)
                    if rr is None:
                        continue
                    hv = rr.get(hcol, None)
                    lv = llm_by_id[_id].get(lcol, None)
                    if not isinstance(hv, str) or not nt(hv):
                        continue
                    if not isinstance(lv, str) or not nt(lv):
                        continue
                    hv = nt(hv)
                    lv = nt(lv)
                    if is_multi_q:
                        rows_multi.append((set(split_multi(hv, hcol)), set(split_multi(lv, hcol))))
                    else:
                        rows_single.append((hv, lv))
                if is_multi_q:
                    m = agreement_metrics_multi(rows_multi)
                else:
                    m = agreement_metrics_single(rows_single)
                coder_stats.append({
                    'track': track.name,
                    'country': country,
                    'coder': coder,
                    'workbook': fp.name,
                    'question': q,
                    'type': 'multi' if is_multi_q else 'single',
                    **m,
                    'shared_ids': len(ids),
                })

        overall_records.append({
            'track': track.name,
            'country': country,
            'total_unique_ids': len(id_to_rows),
            'intercoder_overlap_ids': len(overlap_ids),
            'master_non_overlap_ids': len(non_overlap_ids),
            'llm_ids': len(llm_df),
            'shared_ids_consensus_vs_llm': len(shared_ids),
        })

    # save master workbook in Human Codebooks folder
    master_path = HUMAN_DIR / f'Master {track.name.title()} Codebook - Excluding Intercoder Reliability.xlsx'
    master_wb.save(master_path)

    return overall_records, q_stats, coder_stats, master_path


def main():
    all_overall, all_qstats, all_coderstats = [], [], []
    master_paths = []

    for tr in TRACKS:
        o, q, c, mp = build_master_and_agreement(tr)
        all_overall.extend(o)
        all_qstats.extend(q)
        all_coderstats.extend(c)
        master_paths.append(str(mp))

    # write outputs
    overall_df = pd.DataFrame(all_overall)
    qstats_df = pd.DataFrame(all_qstats)
    coder_df = pd.DataFrame(all_coderstats)

    stamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
    out_xlsx = OUT_DIR / f'human_llm_agreement_full_stats_{stamp}.xlsx'
    out_json = OUT_DIR / f'human_llm_agreement_full_stats_{stamp}.json'

    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as w:
        overall_df.to_excel(w, sheet_name='overview', index=False)
        qstats_df.to_excel(w, sheet_name='question_stats', index=False)
        coder_df.to_excel(w, sheet_name='coder_stats', index=False)

    payload = {
        'master_codebooks': master_paths,
        'overview': all_overall,
        'question_stats': all_qstats,
        'coder_stats': all_coderstats,
    }
    out_json.write_text(json.dumps(payload, indent=2), encoding='utf-8')

    print('MASTER_CODEBOOKS')
    for p in master_paths:
        print('-', p)
    print('AGREEMENT_XLSX', out_xlsx)
    print('AGREEMENT_JSON', out_json)
    print('OVERVIEW_ROWS', len(all_overall), 'QUESTION_STATS_ROWS', len(all_qstats), 'CODER_STATS_ROWS', len(all_coderstats))


if __name__ == '__main__':
    main()
